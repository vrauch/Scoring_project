import os
os.chdir('C:/Users/kaosh/Desktop/Side-project-ML/sentence_analyze')  # 換路徑使用

import openpyxl
wb = openpyxl.load_workbook('TCB Response.xlsx')     # 開啟 Excel 檔案

#names = wb.sheetnames    # 讀取 Excel 裡所有工作表名稱
s1 = wb['DevOps']        # 取得工作表名稱為「工作表1」的內容
s2 = wb['Criteria']           # 取得開啟試算表後立刻顯示的工作表 ( 範例為工作表 2 )

#print(names)
# 印出 title ( 工作表名稱 )、max_row 最大列數、max_column 最大行數
#print(s1.title, s1.max_row, s1.max_column)
#print(s2.title, s2.max_row, s2.max_column)


#print(s1.cell(1, 11).value)   # 等同取出 A1 的內容

def get_values(sheet):
    arr = []                      # 第一層串列
    for row in sheet:
        arr2 = []                 # 第二層串列
        for column in row:
            arr2.append(column.value)  # 寫入內容
        arr.append(arr2)
    return arr
#print(get_values(s1))       # 印出工作表 1 所有內容
#print(get_values(s2))       # 印出工作表 2 所有內容

#------------------------------------- Critera ----------------------------------------
criteria_read = s2.iter_rows(min_row=2, min_col=3, max_col=4)
criteria = []
for i in criteria_read:
    tmp = []
    for j in i:
        tmp.append(j.value)
    criteria.append(tmp)
#print(criteria[0][0])
#print(criteria[0][1])

#--------------------------------------Domain -----------------------------------------
import re
def remove_chinese(text):
    result = ''
    for char in text:
        if '\u4e00' <= char <= '\u9fa5':
            continue
        result += char
    return result

domain_question_read = s1.iter_rows(max_row= 1, min_col=10)
domain_answer_read = s1.iter_rows(min_row = 2, min_col=10)
domain_question = []
domain_answer = []
for rows in domain_question_read:
    tmp = []
    for columns in rows:
        tmp.append(re.sub(r"[^\w\s]", "", remove_chinese(columns.value)).replace("\n",""))            
    domain_question.append(tmp)

print(domain_question)
#print(domain_answer)

            
'''
criteria = []
for row in s2:
    tmp = []
    for column in row:
        tmp.append(column.value)
    criteria.append(tmp)

print(criteria[1][2])
print(criteria[1][3])

for i in range(10, len(s1), 2):
    for j in range()
    if s1.cell(1, i).value in criteria:

'''
