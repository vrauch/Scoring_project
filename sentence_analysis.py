import openai
import os

os.chdir('C:/Users/kaosh/Desktop/Side-project-ML/sentence_analyze')  # 換路徑使用

import openpyxl
wb = openpyxl.load_workbook('TCB Response.xlsx')     # 開啟 Excel 檔案

#names = wb.sheetnames    # 讀取 Excel 裡所有工作表名稱
s1 = wb['DevOps1']        # 取得工作表名稱為「工作表1」的內容

s1_read = s1.iter_rows(max_col=2)
Domain = []
for row in s1_read:
    tmp = []
    for column in row:
        tmp.append(column.value)
    Domain.append(tmp)
#print(len(Domain))
#print(Domain[2][1])






# Set your OpenAI API key
openai.api_key = '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'

# The analyze_sentence function sends a request to the OpenAI API with the given sentence and criteria, and returns the decision made by the model.
def analyze_sentence(sentence, criteria):
    response = openai.Completion.create(
        engine="gpt-3.5-turbo-instruct",
        prompt=f"Sentence: {sentence}\nCriteria: {criteria}\nDecision:",
        temperature=0.9,
        max_tokens=100,
        stop=["\n"]
    )
    decision = response.choices[0].text.strip()
    return decision

# The main function prompts the user to input the sentence and criteria they want to analyze, calls analyze_sentence with these inputs, and prints out the decision.
def main():
    result = []
    for i in range(len(Domain)):
        if Domain[i][0] == "":
            result.append("")
        else: 
            sentence = Domain[i][0]
            criteria = "What percentage is the sentences meet the criteria: " + Domain[i][1]
            decision = analyze_sentence(sentence, criteria)
            result.append(str(f'{decision}'))
    print(result)
    
    wb = openpyxl.load_workbook('TCB Response.xlsx')     # 開啟 Excel 檔案
    s3 = wb.create_sheet("Result", 0) 
    for w in result:
        s3.append(w)
    wb.save('TCB Response.xlsx')
    
    '''
    sentence = "At present, the Bank is developing a 'technical mid-end' design method. Systems in specific fields are required to use specified technical standards for design and development. This method allows IT personnel to be more focused and familiar with the corresponding technical standards and methods. Contribute to project development and maintenance operations."
    #sentence = input("Enter the sentence you want to analyze: ")
    criteria = " Contribution to IT Projects: Examples of how reference architectures ensure security, sustainability, and deployment efficiency.<br>Update Process: The process for keeping architectures updated with technological and cost objectives, including frequency and criteria for updates. "
    #criteria = input("Enter the criteria for analysis: ")
    criteria_b = "What percentage is the sentences meet the criteria: " + criteria
    decision = analyze_sentence(sentence, criteria_b)
    #print(f"The decision for the sentence '{sentence}' based on the criteria '{criteria}' is: {decision}")
    print(f'{decision}')
    '''

if __name__ == "__main__":
    main()
