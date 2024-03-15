import os

os.environ['CURL_CA_BUNDLE'] = ''
from sentence_transformers import SentenceTransformer
from sklearn.metrics.pairwise import cosine_similarity

os.chdir('C:/Users/kaosh/Desktop')  # Change to your own path
import openpyxl
wb = openpyxl.load_workbook('1.xlsx', data_only=True)     # Open Excel file
s1 = wb["Sheet1"]        # Get the worksheet content

# Store all Capabilities in an array
col1_read = s1.iter_rows(max_col=1)
col1 = []
for row in col1_read:
    tmp = []
    for column in row:
        tmp.append(column.value)
    col1.append(tmp)

# Store all questions in an array
col2_read = s1.iter_rows(min_col=2, max_col=2)
col2 = []
for row in col2_read:
    tmp = []
    for column in row:
        tmp.append(column.value)
    col2.append(tmp)
result = [[] for _ in range(len(col1))]
# Load a pre-trained model
model = SentenceTransformer('all-MiniLM-L6-v2')
for i in range(len(col1)):
    if col1[i][0] == None:
        result[i].append("")
    else: 
        if col1[i][0].strip("/") == "":
            result[i].append("")
        else:
            corpus = [col1[i][0]]
            keywords = [col2[i][0]]
            corpus_embeddings = model.encode(corpus)
            keywords_embeddings = model.encode(keywords)
            # Generate embeddings for the corpus and keywords
            # Calculate the average keyword embedding (representative of the set of keywords)
            avg_keywords_embedding = keywords_embeddings.mean(axis=0).reshape(1, -1)
            # Compute the semantic similarity between the corpus and the average keyword embedding
            similarity_scores = cosine_similarity(corpus_embeddings, avg_keywords_embedding)
            print(f"Semantic Similarity Score: {similarity_scores[0][0]}")
            result[i].append(str(similarity_scores[0][0]))

# Save the results in excel '...'
excel = openpyxl.Workbook('Result.xlsx')     
s3 = excel.create_sheet() 
for w in result:
    s3.append(w)
excel.save('Result.xlsx')